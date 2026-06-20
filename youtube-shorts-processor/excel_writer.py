"""Excel logging with embedded screenshots (openpyxl).

Each processed Short becomes one row:
    Timestamp | URL | Screenshot Path | Summary | (embedded image)

Important design note (avoids a real openpyxl pitfall):
    openpyxl's ``Image._data()`` reads the image stream and then *closes* it.
    For images created from a **file path string** that's harmless — openpyxl
    reopens the file fresh on every save. But for images **loaded from an
    existing .xlsx** (whose ref is an in-memory ``BytesIO``), the stream is
    closed after the first save and any subsequent save raises
    "I/O operation on closed file". Because we save incrementally after every
    video, that bug bites on the 2nd row whenever the workbook already existed.

    The fix: we never let openpyxl round-trip embedded images. On startup we
    read existing rows as **text only** (including the stored screenshot path)
    and re-embed every image from its PNG file via a string path. All images
    therefore always have path refs and incremental saves are safe.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config

logger = logging.getLogger(__name__)

HEADERS = ["Timestamp", "URL", "Screenshot Path", "Summary", "Screenshot"]

# Column index (1-based) of the embedded-image column.
_IMAGE_COL = HEADERS.index("Screenshot") + 1

# Conversion constant: Excel row height is in points; ~0.75 pt per pixel.
_PX_TO_POINTS = 0.75
# Excel column width unit ≈ width of one character; ~7 px per unit.
_PX_TO_COL_WIDTH = 1 / 7.0


class ExcelLogger:
    """Append rows (with embedded screenshots) to the summary workbook.

    Holds a single in-memory workbook. Every embedded image is created from a
    screenshot file path (never a loaded stream), so saving repeatedly is safe.
    """

    def __init__(self, excel_path: Path | str | None = None) -> None:
        # Resolve config.EXCEL_FILE at instantiation time (not import time) so
        # the path can be overridden via config before constructing.
        self.excel_path = Path(excel_path) if excel_path else Path(config.EXCEL_FILE)
        self.workbook: Workbook = Workbook()
        self.sheet: Worksheet = self.workbook.active
        self.sheet.title = config.EXCEL_SHEET_NAME
        self._row_cursor = 2  # first data row (row 1 is the header)
        self._write_header()
        self._restore_existing_rows()

    # -- setup --------------------------------------------------------------
    def _write_header(self) -> None:
        header_fill = PatternFill("solid", fgColor="305496")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, title in enumerate(HEADERS, start=1):
            cell = self.sheet.cell(row=1, column=col_idx, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        self.sheet.column_dimensions[get_column_letter(1)].width = 20  # Timestamp
        self.sheet.column_dimensions[get_column_letter(2)].width = 40  # URL
        self.sheet.column_dimensions[get_column_letter(3)].width = 35  # Path
        self.sheet.column_dimensions[get_column_letter(4)].width = 60  # Summary
        self.sheet.freeze_panes = "A2"

    def _restore_existing_rows(self) -> None:
        """Re-load prior rows from an existing workbook as text, then re-embed
        their screenshots from disk. Images are NOT round-tripped through
        openpyxl (see module docstring)."""
        if not self.excel_path.exists():
            return
        try:
            wb = load_workbook(self.excel_path, read_only=True)
            src = (
                wb[config.EXCEL_SHEET_NAME]
                if config.EXCEL_SHEET_NAME in wb.sheetnames
                else wb.active
            )
            existing = list(src.iter_rows(min_row=2, values_only=True))
            wb.close()
        except Exception as exc:
            logger.warning(
                "Could not read existing workbook %s (%s); starting fresh.",
                self.excel_path.name,
                exc,
            )
            return

        restored = 0
        for vals in existing:
            if not vals or all(v is None for v in vals):
                continue
            ts = vals[0] if len(vals) > 0 and vals[0] is not None else ""
            url = vals[1] if len(vals) > 1 and vals[1] is not None else ""
            path = vals[2] if len(vals) > 2 and vals[2] is not None else ""
            summary = vals[3] if len(vals) > 3 and vals[3] is not None else ""
            if not (ts or url or summary):
                continue
            self._add_row(str(ts), str(url), str(path), str(summary))
            restored += 1
        if restored:
            logger.info("Restored %d existing row(s) from %s", restored, self.excel_path.name)

    # -- image sizing -------------------------------------------------------
    @staticmethod
    def _scaled_dimensions(image_path: Path) -> tuple[int, int]:
        """Return (width_px, height_px) scaled to the configured height."""
        target_h = config.EMBED_IMAGE_HEIGHT_PX
        try:
            from PIL import Image as PILImage

            with PILImage.open(image_path) as im:
                orig_w, orig_h = im.size
            if orig_h:
                scale = target_h / float(orig_h)
                return max(1, int(orig_w * scale)), target_h
        except Exception as exc:
            logger.debug("Could not read image size for %s: %s", image_path, exc)
        # Fallback to a portrait-ish Shorts aspect ratio (9:16).
        return int(target_h * 9 / 16), target_h

    # -- row writing --------------------------------------------------------
    def append_row(
        self,
        url: str,
        screenshot_path: Path | str | None,
        summary: str,
        timestamp: str | None = None,
    ) -> int:
        """Append a single result row and embed its screenshot.

        Returns the 1-based row index that was written.
        """
        ts = timestamp or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        path_str = str(screenshot_path) if screenshot_path else ""
        return self._add_row(ts, url or "", path_str, summary or "")

    def _add_row(self, ts: str, url: str, path_str: str, summary: str) -> int:
        row = self._row_cursor
        wrap = Alignment(vertical="center", wrap_text=True)

        self.sheet.cell(row=row, column=1, value=ts).alignment = wrap
        url_cell = self.sheet.cell(row=row, column=2, value=url)
        url_cell.alignment = wrap
        if url:
            try:
                url_cell.hyperlink = url
                url_cell.font = Font(color="0563C1", underline="single")
            except Exception:
                pass
        self.sheet.cell(row=row, column=3, value=path_str).alignment = wrap
        self.sheet.cell(row=row, column=4, value=summary).alignment = wrap

        self._embed_image(row, path_str)
        self._row_cursor += 1
        return row

    def _embed_image(self, row: int, path_str: str) -> None:
        """Embed and size the screenshot (by file path) in the image column."""
        # Default row height even when there is no image (keeps text readable).
        self.sheet.row_dimensions[row].height = max(
            18, config.EMBED_IMAGE_HEIGHT_PX * _PX_TO_POINTS
        )
        if not path_str:
            return
        path = Path(path_str)
        if not path.exists():
            logger.warning("Screenshot not found, skipping embed: %s", path)
            return
        try:
            width_px, height_px = self._scaled_dimensions(path)
            # Pass a *string path* so openpyxl reopens the file on every save;
            # this avoids the closed-stream bug with incremental saves.
            img = XLImage(str(path))
            img.width = width_px
            img.height = height_px

            col_letter = get_column_letter(_IMAGE_COL)
            self.sheet.add_image(img, f"{col_letter}{row}")

            needed_width = width_px * _PX_TO_COL_WIDTH + 2
            current = self.sheet.column_dimensions[col_letter].width or 0
            if needed_width > current:
                self.sheet.column_dimensions[col_letter].width = needed_width
        except Exception as exc:
            logger.error("Failed to embed image for row %d: %s", row, exc)

    # -- persistence --------------------------------------------------------
    def save(self) -> None:
        """Persist the workbook to disk."""
        try:
            self.excel_path.parent.mkdir(parents=True, exist_ok=True)
            self.workbook.save(self.excel_path)
            logger.info("Saved workbook -> %s", self.excel_path)
        except PermissionError:
            logger.error(
                "Could not save %s. Is it open in Excel? Close it and retry.",
                self.excel_path,
            )
            raise
